"""
Excel (.xlsx) text extractor.

Banking Excel files often contain rate tables, exception matrices,
product parameters, and compliance checklists. Each sheet is treated
as a separate logical section.

Rows are serialized as "Column: Value" pairs to preserve meaning
when embedded, rather than raw CSV which loses column context.
"""

from pathlib import Path
from typing import Any

import openpyxl

from utils.logger import get_logger

logger = get_logger(__name__)


def extract(file_path: str | Path) -> list[dict[str, Any]]:
    """
    Extract text from an Excel (.xlsx) file, sheet by sheet.

    Args:
        file_path: Path to the Excel file.

    Returns:
        List of sheet-level dicts with serialized row text and metadata.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        logger.error("XLSX file not found: %s", file_path)
        return []

    sections: list[dict[str, Any]] = []

    try:
        workbook = openpyxl.load_workbook(str(file_path), data_only=True)
        logger.info(
            "Opened Excel: %s (%d sheets)", file_path.name, len(workbook.sheetnames)
        )

        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            sheet_text = _serialize_sheet(sheet, sheet_name)

            if sheet_text.strip():
                sections.append({
                    "text":     sheet_text.strip(),
                    "source":   file_path.name,
                    "page":     sheet_index + 1,
                    "section":  sheet_name,
                    "doc_type": "xlsx",
                })

    except Exception as exc:
        logger.error("Failed to extract XLSX %s: %s", file_path.name, exc)
        return []

    logger.info(
        "Extracted %d sheet(s) from %s", len(sections), file_path.name
    )
    return sections


def _serialize_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str) -> str:
    """
    Serialize a worksheet into readable text.

    Header row (row 1) is used as column names. Subsequent rows are formatted
    as "ColumnName: Value" pairs, one per line, rows separated by blank lines.
    This format preserves column context during embedding.
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""

    # Extract headers from first row; fall back to Column_N if header is empty
    headers = [
        str(cell).strip() if cell is not None else f"Column_{i+1}"
        for i, cell in enumerate(rows[0])
    ]

    lines = [f"Sheet: {sheet_name}", ""]

    for row in rows[1:]:
        row_parts = []
        for header, value in zip(headers, row):
            if value is not None and str(value).strip():
                row_parts.append(f"{header}: {value}")
        if row_parts:
            lines.append(" | ".join(row_parts))

    return "\n".join(lines)